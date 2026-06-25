"""Semantic intermediate normalizers for spreadsheet, PDF, and Macro 96 Excel.

Writes structured intermediate artifacts only — never raw/candidate/final/kg_*.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

_MAX_PREVIEW_ROWS = 20

# Macro 96 standard pool column signatures (Brain volume list.xlsx)
_MACRO_POOL_INDEX_HEADERS = {"id #", "id", "index", "pool_id", "pool_index"}
_MACRO_EN_NAME_HEADERS = {"brain structure", "brain_structure", "name_en", "en_name"}
_MACRO_CN_NAME_HEADERS = {"脑区中文名称", "cn_name", "name_cn", "中文名称", "中文脑区名称"}

NORMALIZER_KEY_SPREADSHEET = "spreadsheet_workbook_v1"
NORMALIZER_KEY_MACRO_REGION = "macro_region_table_v1"
NORMALIZER_KEY_PDF = "pdf_metadata_v1"
NORMALIZER_KEY_DOCUMENT = "document_metadata_v1"

ARTIFACT_KIND_PRIORITY = (
    "macro_region_table",
    "label_table",
    "spreadsheet_workbook",
    "tabular_data",
    "json_document",
    "text_document",
    "pdf_metadata",
    "document_metadata",
    "ontology_document",
    "image_metadata",
    "nifti_metadata",
    "connectivity_matrix_metadata",
    "binary_metadata",
    "unsupported",
)


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def is_brain_volume_list_filename(filename: str) -> bool:
    name = filename.lower().replace("_", " ")
    return "brain volume" in name or name.startswith("brain volume list")


def detect_macro_96_column_map(headers: list[str]) -> dict[str, str] | None:
    """Return mapping internal_field -> original_header if sheet matches Macro 96 layout."""
    if not headers:
        return None
    idx_map: dict[str, str] = {}
    for h in headers:
        nh = _norm_header(h)
        if nh in _MACRO_POOL_INDEX_HEADERS and "region_index" not in idx_map:
            idx_map["region_index"] = h
        elif nh in _MACRO_EN_NAME_HEADERS and "en_name" not in idx_map:
            idx_map["en_name"] = h
        elif nh in _MACRO_CN_NAME_HEADERS and "cn_name" not in idx_map:
            idx_map["cn_name"] = h
    if {"region_index", "en_name", "cn_name"}.issubset(idx_map.keys()):
        return idx_map
    return None


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _read_xlsx_sheets(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    """Return (sheets_data, warnings). Each sheet: name, headers, rows, row_count."""
    warnings: list[str] = []
    try:
        import openpyxl
    except ImportError as exc:
        warnings.append(f"openpyxl not available: {exc}")
        return [], warnings

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        warnings.append(f"Excel open error: {exc}")
        return [], warnings

    sheets: list[dict[str, Any]] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                sheets.append({
                    "sheet_name": sheet_name,
                    "row_count": 0,
                    "column_count": 0,
                    "columns": [],
                    "rows_preview": [],
                })
                continue

            headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
            data_rows: list[list[Any]] = []
            for row in rows_iter:
                if row is None or all(c is None or str(c).strip() == "" for c in row):
                    continue
                data_rows.append(list(row))

            dict_rows: list[dict[str, Any]] = []
            for row in data_rows:
                item: dict[str, Any] = {}
                for i, col in enumerate(headers):
                    if i < len(row):
                        item[col] = _cell_value(row[i])
                dict_rows.append(item)

            sheets.append({
                "sheet_name": sheet_name,
                "row_count": len(dict_rows),
                "column_count": len(headers),
                "columns": headers,
                "rows_preview": dict_rows[:_MAX_PREVIEW_ROWS],
                "_all_rows": dict_rows,
            })
    finally:
        wb.close()

    return sheets, warnings


def _build_macro_region_table(
    sheet: dict[str, Any],
    *,
    provenance: dict[str, Any],
    filename: str,
) -> dict[str, Any] | None:
    col_map = detect_macro_96_column_map(sheet["columns"])
    if col_map is None:
        return None

    all_rows = sheet.get("_all_rows") or sheet.get("rows_preview") or []
    mapped: list[dict[str, Any]] = []
    warnings: list[str] = []

    for raw in all_rows:
        region_index = _cell_value(raw.get(col_map["region_index"]))
        en_name = raw.get(col_map["en_name"])
        cn_name = raw.get(col_map["cn_name"])
        if region_index is None and not en_name:
            continue
        try:
            idx_int = int(region_index) if region_index is not None else None
        except (TypeError, ValueError):
            idx_int = None
            warnings.append(f"Non-integer region_index: {region_index!r}")
        if not en_name:
            warnings.append(f"Empty en_name at index {region_index}")
        mapped.append({
            "region_index": idx_int,
            "en_name": str(en_name).strip() if en_name is not None else "",
            "cn_name": str(cn_name).strip() if cn_name is not None else "",
        })

    preview_rows = mapped[:_MAX_PREVIEW_ROWS]
    return {
        "artifact_kind": "macro_region_table",
        "source_format": "xlsx",
        "row_count": len(mapped),
        "content_jsonb": {
            "schema": "macro_region_table_v1",
            "source_format": "xlsx",
            "columns": ["region_index", "en_name", "cn_name"],
            "rows": mapped,
            "row_count": len(mapped),
            "source_sheet": sheet["sheet_name"],
        },
        "preview_jsonb": {
            "preview_limit": _MAX_PREVIEW_ROWS,
            "total_rows": len(mapped),
            "rows_preview": preview_rows,
            "columns": ["region_index", "en_name", "cn_name"],
        },
        "metadata_jsonb": {
            "detected_from": filename,
            "matched_columns": [col_map["region_index"], col_map["en_name"], col_map["cn_name"]],
            "purpose": "macro_96_region_pool_intermediate",
            "note": "This is an intermediate artifact only. It does not create macro_96_pool records.",
            "source_sheet": sheet["sheet_name"],
            **provenance,
        },
        "warnings_jsonb": warnings,
    }


def normalize_spreadsheet_workbook(
    path: Path,
    ext: str,
    provenance: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any] | None]:
    """Parse .xlsx/.xls into spreadsheet_workbook; optionally macro_region_table."""
    source_format = "xlsx" if ext == ".xlsx" else "xls"
    filename = provenance.get("original_filename") or path.name
    sheets_raw, open_warnings = _read_xlsx_sheets(path)

    if not sheets_raw:
        return {
            "artifact_kind": "spreadsheet_workbook",
            "source_format": source_format,
            "row_count": None,
            "content_jsonb": None,
            "preview_jsonb": None,
            "metadata_jsonb": {
                "original_filename": filename,
                "sheet_count": 0,
                "detected_table_kinds": [],
                **provenance,
            },
            "warnings_jsonb": open_warnings or ["Could not parse spreadsheet structure"],
        }, None

    # Strip internal _all_rows from content export
    sheets_content: list[dict[str, Any]] = []
    detected_kinds: list[str] = []
    macro_artifact: dict[str, Any] | None = None

    for s in sheets_raw:
        sheets_content.append({
            "sheet_name": s["sheet_name"],
            "row_count": s["row_count"],
            "column_count": s["column_count"],
            "columns": s["columns"],
            "rows_preview": s["rows_preview"],
        })
        if detect_macro_96_column_map(s["columns"]) is not None:
            detected_kinds.append("macro_region_table")
            if macro_artifact is None:
                macro_artifact = _build_macro_region_table(s, provenance=provenance, filename=filename)

    primary = sheets_content[0] if sheets_content else None
    total_rows = sum(s["row_count"] for s in sheets_content)

    workbook_artifact: dict[str, Any] = {
        "artifact_kind": "spreadsheet_workbook",
        "source_format": source_format,
        "row_count": total_rows,
        "content_jsonb": {
            "schema": "spreadsheet_workbook_v1",
            "source_format": source_format,
            "sheets": sheets_content,
        },
        "preview_jsonb": {
            "primary_sheet": primary["sheet_name"] if primary else None,
            "rows_preview": primary["rows_preview"] if primary else [],
            "columns": primary["columns"] if primary else [],
            "preview_limit": _MAX_PREVIEW_ROWS,
        },
        "metadata_jsonb": {
            "original_filename": filename,
            "sheet_count": len(sheets_content),
            "detected_table_kinds": detected_kinds,
            **provenance,
        },
        "warnings_jsonb": open_warnings,
    }

    if is_brain_volume_list_filename(filename) and macro_artifact is None:
        workbook_artifact["warnings_jsonb"] = list(workbook_artifact["warnings_jsonb"]) + [
            "Filename suggests Brain volume list but Macro 96 columns were not detected",
        ]

    return workbook_artifact, macro_artifact


def normalize_pdf_metadata(path: Path, provenance: dict[str, Any]) -> dict[str, Any]:
    """PDF metadata only — no OCR / full text extraction in this module."""
    size = path.stat().st_size
    filename = provenance.get("original_filename") or path.name
    metadata: dict[str, Any] = {
        "schema": "pdf_metadata_v1",
        "source_format": "pdf",
        "original_filename": filename,
        "file_size": size,
        "text_extraction": "not_implemented",
        "ocr": "not_implemented",
        "note": "PDF text extraction is not implemented in this module.",
        **provenance,
    }
    warnings: list[str] = []

    try:
        from pypdf import PdfReader  # optional
        reader = PdfReader(str(path))
        metadata["page_count"] = len(reader.pages)
        info = reader.metadata
        if info:
            if info.title:
                metadata["title"] = str(info.title)
            if info.author:
                metadata["author"] = str(info.author)
        if reader.pages:
            try:
                text = reader.pages[0].extract_text() or ""
                metadata["first_page_text_preview"] = text[:2000]
                metadata["text_extraction"] = "first_page_preview_only"
            except Exception as exc:
                warnings.append(f"First page text preview failed: {exc}")
    except ImportError:
        metadata["pypdf_available"] = False
    except Exception as exc:
        warnings.append(f"PDF read error: {exc}")

    return {
        "artifact_kind": "pdf_metadata",
        "source_format": "pdf",
        "row_count": metadata.get("page_count"),
        "content_jsonb": None,
        "preview_jsonb": {
            "page_count": metadata.get("page_count"),
            "title": metadata.get("title"),
            "first_page_text_preview": metadata.get("first_page_text_preview"),
        },
        "metadata_jsonb": metadata,
        "warnings_jsonb": warnings,
    }


def pick_primary_artifact_kind(kinds: list[str]) -> str | None:
    if not kinds:
        return None
    kind_set = set(kinds)
    for k in ARTIFACT_KIND_PRIORITY:
        if k in kind_set:
            return k
    return kinds[0]
