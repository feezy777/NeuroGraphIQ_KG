from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Tuple

from ..common.id_utils import make_id
from ..common.models import ContentChunk, ParsedDocument

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


def _split_sentences(text: str) -> List[str]:
    if not text:
        return []
    normalized = text.replace("\r\n", "\n")
    for mark in ["\n", ".", "!", "?", ";", "。", "！", "？", "；"]:
        normalized = normalized.replace(mark, "|||")
    return [it.strip() for it in normalized.split("|||") if it.strip()]


def _read_text_file(path: Path) -> Tuple[str, Dict[str, Any]]:
    raw = path.read_text(encoding="utf-8", errors="ignore")
    meta: Dict[str, Any] = {"parser": "text_reader"}
    if path.suffix.lower() in {".json", ".jsonl"}:
        try:
            row = raw.splitlines()[0] if path.suffix.lower() == ".jsonl" else raw
            payload = json.loads(row)
            meta["json_kind"] = "dict" if isinstance(payload, dict) else "list"
        except Exception:
            meta["json_kind"] = "invalid_json"
    return raw, meta


def _read_pdf_file(path: Path) -> Tuple[str, Dict[str, Any]]:
    if PdfReader is None:
        return "", {"parser": "pdf_reader", "warning": "pypdf_not_installed"}
    reader = PdfReader(str(path))
    pages: List[str] = []
    for i, page in enumerate(reader.pages):
        txt = (page.extract_text() or "").strip()
        if txt:
            pages.append(f"[page={i + 1}] {txt}")
    return "\n".join(pages), {"parser": "pdf_reader", "page_count": len(reader.pages), "extracted_pages": len(pages)}


def _read_xlsx_file(path: Path) -> Tuple[str, List[Dict[str, Any]], Dict[str, Any]]:
    if load_workbook is None:
        return "", [], {"parser": "xlsx_reader", "warning": "openpyxl_not_installed"}
    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows: List[str] = []
    table_cells: List[Dict[str, Any]] = []
    max_rows = 80
    for ridx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        values = [str(it).strip() for it in row if it is not None and str(it).strip()]
        if not values:
            continue
        rows.append(" | ".join(values))
        for cidx, val in enumerate(values, start=1):
            table_cells.append({"sheet": sheet.title, "row": ridx, "col": cidx, "value": val})
    wb.close()
    return "\n".join(rows), table_cells, {"parser": "xlsx_reader", "sheet": sheet.title, "row_count": len(rows)}


class ParsingService:
    def parse_file(self, file_payload: Dict[str, Any]) -> Dict[str, Any]:
        file_id = file_payload["file_id"]
        file_type = (file_payload.get("file_type") or "unknown").lower()
        path = Path(file_payload["path"])

        raw_text = ""
        table_cells: List[Dict[str, Any]] = []
        metadata: Dict[str, Any] = {}

        if file_type in {"txt", "md", "json", "jsonl", "csv", "tsv"}:
            raw_text, metadata = _read_text_file(path)
        elif file_type in {"pdf"}:
            raw_text, metadata = _read_pdf_file(path)
        elif file_type in {"xlsx", "xls"}:
            raw_text, table_cells, metadata = _read_xlsx_file(path)
        else:
            raw_text, metadata = _read_text_file(path)
            metadata["warning"] = f"fallback_reader_for_{file_type}"

        paragraphs = [it.strip() for it in raw_text.splitlines() if it.strip()]
        sentences = _split_sentences(raw_text)[:500]

        doc = ParsedDocument(
            parsed_document_id=make_id("pd"),
            file_id=file_id,
            parse_status="parsed_success" if raw_text or table_cells else "parsed_failed",
            file_type=file_type,
            title=Path(file_payload.get("filename", path.name)).stem,
            source=file_payload.get("filename", path.name),
            authors=[],
            year=None,
            doi="",
            page_range="",
            raw_text=raw_text,
            metadata_json=metadata,
            paragraphs=paragraphs[:500],
            sentences=sentences,
            table_cells=table_cells[:1000],
            figure_captions=[],
            heading_levels=[],
            ocr_blocks=[],
            parser_name=metadata.get("parser", f"{file_type}_parser"),
            parser_version="v1",
        )

        chunks: List[Dict[str, Any]] = []
        for idx, para in enumerate(doc.paragraphs[:400], start=1):
            chunk = ContentChunk(
                chunk_id=make_id("chk"),
                file_id=file_id,
                chunk_type="paragraph",
                chunk_index=idx,
                text_content=para,
                page_no=None,
                source_ref=f"paragraph:{idx}",
                extra_json={"from": "paragraph"},
            )
            chunks.append(
                {
                    "chunk_id": chunk.chunk_id,
                    "file_id": chunk.file_id,
                    "chunk_type": chunk.chunk_type,
                    "chunk_index": chunk.chunk_index,
                    "text_content": chunk.text_content,
                    "page_no": chunk.page_no,
                    "source_ref": chunk.source_ref,
                    "extra_json": chunk.extra_json,
                }
            )

        for idx, cell in enumerate(doc.table_cells[:200], start=1):
            chunk = ContentChunk(
                chunk_id=make_id("chk"),
                file_id=file_id,
                chunk_type="table_cell",
                chunk_index=idx,
                text_content=str(cell.get("value", "")),
                page_no=None,
                source_ref=f"{cell.get('sheet', 'Sheet1')}:{cell.get('row', 0)}:{cell.get('col', 0)}",
                extra_json={"from": "table_cell", **cell},
            )
            chunks.append(
                {
                    "chunk_id": chunk.chunk_id,
                    "file_id": chunk.file_id,
                    "chunk_type": chunk.chunk_type,
                    "chunk_index": chunk.chunk_index,
                    "text_content": chunk.text_content,
                    "page_no": chunk.page_no,
                    "source_ref": chunk.source_ref,
                    "extra_json": chunk.extra_json,
                }
            )

        return {"document": doc, "chunks": chunks}
